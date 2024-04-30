import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index = 1, title = 'Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'
ws['A1'].font = Font(name = 'Times New Roman', size = 24, bold = True)

headerfont = Font(name = 'Times New Roman', size = 24, bold = True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')
ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

#to unmerge cells
ws['A8'] = 'Total'
ws['A8'].font = Font(size=16, bold=True)

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

write_ws = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

for row_index in range(1, read_ws.max_row + 1):
    for col_index in range(1, read_ws.max_column + 1):
        cell_value = read_ws.cell(row_index, col_index).value
        write_ws.cell(row_index, col_index, cell_value)


write_ws['A43'] = 'Amt Sold Total'
write_ws['A43'].font = Font(size=16, bold=True)
write_ws['B43'] = '=SUM(C2:C41)'

write_ws['A44'] = 'Amt Sold Average'
write_ws['A44'].font = Font(size=16, bold=True)
write_ws['B44'] = '=AVERAGE(C2:C41)'

write_ws['A45'] = 'Grand Total Total'
write_ws['A45'].font = Font(size=16, bold=True)
write_ws['B45'] = '=SUM(D2:D41)'

write_ws['A46'] = 'Total Average'
write_ws['A46'].font = Font(size=16, bold=True)
write_ws['B46'] = '=AVERAGE(D2:D41)'


read_wb.save('ProduceReport.xlsx')
wb.save("PythontoExcel.xlsx")
